import os
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

print("⏳ جاري البحث عن ملفات Excel في المجلد ودمجها...")

current_folder = os.path.dirname(os.path.abspath(__file__))
target_folder = current_folder 
output_filename = 'merged_branches_summary.xlsx'
output_path = os.path.join(current_folder, output_filename)

all_dfs = []

for filename in os.listdir(target_folder):
    if filename.endswith(('.xlsx', '.xls')) and not filename.startswith('~$') and filename != output_filename:
        file_path = os.path.join(target_folder, filename)
        
        try:
            df = pd.read_excel(file_path)
            all_dfs.append(df)
            print(f"✅ تم قراءة الملف: {filename}")
        except Exception as e:
            print(f"⚠️ تعذر قراءة الملف {filename}: {e}")

if all_dfs:
    all_branches = pd.concat(all_dfs, ignore_index=True)

    final_summary = all_branches.groupby('Product', as_index=False).sum()

    total_row = pd.DataFrame([{
        'Product': '⚠️ TOTAL SALES',
        'Units Sold': final_summary['Units Sold'].sum(),
        'Revenue ($)': final_summary['Revenue ($)'].sum()
    }])
    final_summary = pd.concat([final_summary, total_row], ignore_index=True)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        final_summary.to_excel(writer, index=False, sheet_name='Sales Analytics')
        workbook = writer.book
        worksheet = writer.sheets['Sales Analytics']
        
        worksheet.sheet_view.showGridLines = True

        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # كحلي داكن
        total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # أخضر فاتح
        
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        data_font = Font(name="Segoe UI", size=10, color="333333")
        total_font = Font(name="Segoe UI", size=11, bold=True, color="375623")

        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
        )

        for col_num in range(1, 4):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in range(2, len(final_summary) + 2):
            is_total_row = (row == len(final_summary) + 1)
            for col in range(1, 4):
                cell = worksheet.cell(row=row, column=col)
                cell.border = thin_border
                
                if is_total_row:
                    cell.fill = total_fill
                    cell.font = total_font
                else:
                    cell.font = data_font

                if col == 1:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                elif col == 2:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.number_format = '#,##0'
                elif col == 3:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.number_format = '"$"#,##0.00'

        # ضبط عرض الأعمدة
        for col_idx in range(1, 4):
            max_len = max(len(str(worksheet.cell(row=r, column=col_idx).value or '')) for r in range(1, len(final_summary) + 2))
            col_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[col_letter].width = max(max_len + 5, 18)

    print(f"\n✨ نجاح باهر! تم دمج {len(all_dfs)} ملفات وتوليد التقرير المالي الموحد.")
    print(f"📁 تجد الملف الجاهز باسم: {output_path}")
else:
    print("❌ لم يتم العثور على أي ملفات Excel مطابقة داخل المجلد المحدد.")