import json
import os
import urllib.parse
import urllib.request
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd

url = "https://itunes.apple.com/search?term=programming+coding&media=ebook&limit=40"
headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}

book_titles = []
book_authors = []
book_years = []
book_prices = []
book_currencies = []

print(
    "⏳ جاري الاتصال بسيرفر Apple Books الرسمي لسحب الكتب وأسعارها الحقيقية..."
)

try:
    req = urllib.request.Request(url, headers=headers)
    with urllib.request.urlopen(req) as response:
        raw_data = response.read().decode("utf-8")

    data = json.loads(raw_data)
    books = data.get("results", [])

    for book in books:
        
        title = book.get("trackName", "Unknown Title")

       
        author = book.get("artistName", "Unknown Author")

        release_date = book.get("releaseDate", "N/A")
        pub_year = release_date[:4] if len(release_date) >= 4 else "N/A"

        price = book.get("trackPrice", book.get("price", 0.0))
        currency = book.get("currency", "USD")

        book_titles.append(title)
        book_authors.append(author)
        book_years.append(pub_year)
        book_prices.append(float(price) if price else 0.0)
        book_currencies.append(currency)

except Exception as e:
    print(f"❌ حدث خطأ أثناء الاتصال: {e}")

if book_titles:
    data = {
        "📚 Book Title": book_titles,
        "✍️ Author Name": book_authors,
        "📅 Release Year": book_years,
        "💵 Original Price ($)": book_prices,
        "🔤 Currency": book_currencies,
    }
    df = pd.DataFrame(data)

    current_folder = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(
        current_folder, "apple_books_real_prices_report.xlsx"
    )

    print("🎨 جاري تطبيق التنسيقات الملكية وإنشاء ملف Excel...")

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Programming Books")
        workbook = writer.book
        worksheet = writer.sheets["Programming Books"]

        header_fill = PatternFill(
            start_color="1F4E78", end_color="1F4E78", fill_type="solid"
        )
        header_font = Font(name="Segoe UI", size=12, bold=True, color="FFFFFF")
        data_font = Font(name="Segoe UI", size=10, bold=False, color="333333")

        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )

        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")

        for col_num, column in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

        for row in range(2, len(df) + 2):
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.font = data_font
                cell.border = thin_border

                if col in [1, 2]:
                    cell.alignment = left_align
                elif col in [3, 5]:
                    cell.alignment = center_align
                elif col == 4:
                    cell.alignment = right_align
                    cell.number_format = (
                        '"$"#,##0.00' 
                    )

        for col in worksheet.columns:
            max_len = max(
                len(
                    f"${cell.value:,.2f}"
                    if isinstance(cell.value, (int, float))
                    and col[0].column == 4
                    else str(cell.value or "")
                )
                for cell in col
            )
            col_letter = get_column_letter(col[0].column)
            worksheet.column_dimensions[col_letter].width = max(
                max_len + 5, 15
            )

    print(
        f"✨ نجاح مكتمل وباهر! تم سحب {len(book_titles)} كتاباً برمجياً بأسعارهم الحقيقية المعتمدة من Apple."
    )
    print(
        f"📁 تجد ملف Excel الاحترافي جاهزاً في مجلدك باسم:\n{excel_path}"
    )
else:
    print(
        "❌ لم يتم جمع أي بيانات، يرجى التحقق من اتصال الإنترنت وأعد المحاولة."
    )